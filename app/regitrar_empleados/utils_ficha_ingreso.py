import io
from openpyxl import load_workbook
from copy import copy
from django.conf import settings
import os
from .models import Empleado

# ------------------------------------------------------------
# Mapeos para convertir valores legibles a códigos numéricos de la ficha
# ------------------------------------------------------------

TIPO_DOC_MAP = {
    'CC': '01', 'CE': '02', 'PT': '04', 'TI': '03',
    'RC': '05', 'PA': '06', 'PE': '07', 'NI': '10',
    'TE': '21', 'DE': '22', 'IE': '23',
}

ESTADO_CIVIL_MAP = {
    'Soltero(a)': '1',
    'Casado(a)': '2',
    'Separado(a)': '4',
    'Unión Libre': '5',
    'Viudo(a)': '3',
    'Religiosa(a)': '6',
}

TIPO_CUENTA_NUM_MAP = {
    '1 - Consignación Cuenta Ahorros': '1',
    '2 - Consignación Cuenta Corriente': '2',
    '3 - Pago con Cheque': '3',
    '4 - Pago en Efectivo': '4',
    '5 - Otra Forma de Pago': '5',
}

TIPO_LIQUIDACION_NUM_MAP = {
    '0 - NO APLICA': '0',
    '1 - Quincenal': '1',
    '2 - Mensual': '2',
    '3 - Semanal': '3',
    '4 - Catorcenal': '4',
    '5 - Grupo 2': '5',
}

CLASE_SALARIO_NUM_MAP = {
    '1 - Normal': '1',
    '2 - Integral': '2',
}

TIPO_CONTRATO_NUM_MAP = {
    '01 - Termino indefinido': '01',
    '02 - Termino fijo': '02',
    '03 - Termino indefinido sin transp': '03',
    '05 - Termino fijo < 1 año': '05',
    '06 - Honorarios': '06',
    '09 - Aprendiz Sena': '09',
}

METODO_RETENCION_NUM_MAP = {
    'Modalidad 1': '1',
    'Modalidad 2': '2',
}

MODO_LIQUIDACION_NUM_MAP = {
    '0 - Normal': '0',
    '1 - Aprendiz': '1',
    '2 - Sin transporte': '2',
    '3 - Asumido SS': '3',
    '4 - Especial': '4',
    '5 - Liquidación x Hora': '5',
    '6 - Flexibilización': '6',
    '10 - Valor Hora x Clasif': '10',
    '11 - Valor Hora x Puntos': '11',
}

TIPO_COTIZANTE_NUM_MAP = {
    '01 Dependiente': '01',
    '12 aprendiz en etapa lectiva': '12',
    '19 Aprendiz en etapa productiva': '19',
}

# ------------------------------------------------------------
# Funciones auxiliares
# ------------------------------------------------------------
def extraer_numero_antes_guion(valor):
    """Extrae el primer número o código antes de ' - '"""
    if valor and ' - ' in valor:
        return valor.split(' - ')[0].strip()
    return valor if valor else ''

def extraer_codigo_pais(valor):
    """Extrae el código de país (primer número)"""
    if valor and ' - ' in valor:
        return valor.split(' - ')[0].strip()
    return ''

def extraer_codigo_ciudad(valor):
    """
    Extrae el código de ciudad (tercer número) de formato: '057 - 76 - 76863 - VERSALLES'
    """
    if valor and ' - ' in valor:
        partes = valor.split(' - ')
        if len(partes) >= 3:
            return partes[2].strip()
    return ''

def extraer_codigo_centro_costos(valor):
    """Extrae el código antes de ' - ' (ej: C1004)"""
    if valor and ' - ' in valor:
        return valor.split(' - ')[0].strip()
    return valor if valor else ''

def formato_fecha(fecha):
    return fecha.strftime('%d/%m/%Y') if fecha else ''

def sexo_a_numero(sexo):
    if sexo in ['Femenino', 'F', 'FEMENINO', '1']:
        return '1'
    return '2'

def bool_a_01(valor):
    if valor is None:
        return '0'
    return '1' if valor else '0'

def formato_sueldo(sueldo):
    """Devuelve el sueldo como número entero si no tiene decimales"""
    if sueldo is not None:
        return str(int(sueldo)) if sueldo == int(sueldo) else str(sueldo)
    return ''

def es_extranjero_a_numero(empleado):
    """1 Colombiano, 2 Extranjero (basado en campo extranjero)"""
    return '2' if empleado.extranjero else '1'

# ------------------------------------------------------------
# Mapeo columna -> función extractora (recibe empleado y datos_extra)
# ------------------------------------------------------------
def crear_mapeo_columnas(datos_extra_dict):
    def get_extra(e, key, default=''):
        return datos_extra_dict.get(e.id_empleado, {}).get(key, default)

    MAPEO = {
        'B': lambda e, ex: e.codigo_empleado or '',
        'C': lambda e, ex: e.documento or '',
        'D': lambda e, ex: TIPO_DOC_MAP.get(e.tipo_doc, '00'),
        'E': lambda e, ex: e.codigo_alterno or '0',
        'F': lambda e, ex: (e.primer_apellido or '').upper(),
        'G': lambda e, ex: (e.segundo_apellido or '').upper(),
        'H': lambda e, ex: (e.nombre_1 or '').upper(),
        'I': lambda e, ex: (e.nombre_2 or '').upper(),
        'J': lambda e, ex: formato_fecha(e.f_nacimiento),
        'K': lambda e, ex: sexo_a_numero(e.sexo),
        'L': lambda e, ex: '',
        'M': lambda e, ex: '0',
        'N': lambda e, ex: '',
        'O': lambda e, ex: get_extra(e, 'grupo_sanguineo', e.grupo_sanguineo or ''),
        'P': lambda e, ex: get_extra(e, 'factor_rh', e.factor_rh or ''),
        'Q': lambda e, ex: ESTADO_CIVIL_MAP.get(e.estado_civil, '0') if e.estado_civil else '0',
        'R': lambda e, ex: es_extranjero_a_numero(e),            # usamos el campo extranjero
        'S': lambda e, ex: (e.direccion_residencia or '').upper(),
        'T': lambda e, ex: e.telefono_residencia or '',
        'U': lambda e, ex: '',
        'V': lambda e, ex: e.email or '',
        'W': lambda e, ex: e.celular or '',
        'X': lambda e, ex: formato_fecha(e.f_ingreso),
        'Y': lambda e, ex: TIPO_CUENTA_NUM_MAP.get(e.tipo_cuenta, '0'),
        'Z': lambda e, ex: extraer_numero_antes_guion(e.banco),
        'AA': lambda e, ex: e.numero_cuenta or '',
        'AB': lambda e, ex: TIPO_LIQUIDACION_NUM_MAP.get(e.tipo_liquidacion, '0'),
        'AC': lambda e, ex: '2',
        'AD': lambda e, ex: CLASE_SALARIO_NUM_MAP.get(e.clase_salario, '0'),
        'AE': lambda e, ex: '001',
        'AF': lambda e, ex: '001',
        'AG': lambda e, ex: extraer_codigo_centro_costos(e.centro_costos),
        'AH': lambda e, ex: '0' if (e.subcliente or '').upper() == 'NO APLICA' else extraer_numero_antes_guion(e.subcliente),
        'AI': lambda e, ex: '0' if (e.clasificacion_2 or '').upper() == 'NO APLICA' else (e.clasificacion_2 or '0'),
        'AJ': lambda e, ex: '0' if (e.clasificacion_3 or '').upper() == 'NO APLICA' else (e.clasificacion_3 or '0'),
        'AK': lambda e, ex: '0' if (e.clasificacion_4 or '').upper() == 'NO APLICA' else (e.clasificacion_4 or '0'),
        'AL': lambda e, ex: '0' if (e.clasificacion_5 or '').upper() == 'NO APLICA' else (e.clasificacion_5 or '0'),
        'AM': lambda e, ex: '0' if (e.clasificacion_6 or '').upper() == 'NO APLICA' else (e.clasificacion_6 or '0'),
        'AN': lambda e, ex: '0' if (e.clasificacion_7 or '').upper() == 'NO APLICA' else (e.clasificacion_7 or '0'),
        'AO': lambda e, ex: (get_extra(e, 'cargo', e.cargo or '')).upper(),
        'AP': lambda e, ex: TIPO_CONTRATO_NUM_MAP.get(e.tipo_contrato, '00'),
        'AQ': lambda e, ex: '300',                              # Fondo riesgos siempre 300
        'AR': lambda e, ex: '2,436',                            # % riesgos fijo
        'AS': lambda e, ex: extraer_numero_antes_guion(e.afp),
        'AT': lambda e, ex: extraer_numero_antes_guion(e.eps),
        'AU': lambda e, ex: extraer_numero_antes_guion(e.ccf),
        'AV': lambda e, ex: extraer_numero_antes_guion(e.fondo_cesantias),
        'AW': lambda e, ex: METODO_RETENCION_NUM_MAP.get(e.metodo_retencion, '0'),
        'AX': lambda e, ex: '0' if METODO_RETENCION_NUM_MAP.get(e.metodo_retencion, '0') == '1' else str(e.porcentaje_ret) if e.porcentaje_ret is not None else '0',
        'AY': lambda e, ex: formato_sueldo(e.sueldo),           # sueldo sin decimales
        'AZ': lambda e, ex: str(e.vacaciones) if e.vacaciones is not None else '15',
        'BA': lambda e, ex: get_extra(e, 'sabado_habil', e.sabado_habil or '0'),
        'BB': lambda e, ex: '0',
        'BC': lambda e, ex: e.cuenta_gastos or '',
        'BD': lambda e, ex: bool_a_01(e.variable),
        'BE': lambda e, ex: bool_a_01(e.pago_por_dias),
        'BF': lambda e, ex: MODO_LIQUIDACION_NUM_MAP.get(e.modo_liquidacion_conceptos, '0'),
        'BG': lambda e, ex: extraer_codigo_pais(e.ciudad_exp_documento) if e.ciudad_exp_documento else '',
        'BH': lambda e, ex: extraer_codigo_ciudad(e.ciudad_exp_documento) if e.ciudad_exp_documento else '',
        'BI': lambda e, ex: extraer_codigo_pais(e.lugar_nacimiento),
        'BJ': lambda e, ex: extraer_codigo_ciudad(e.lugar_nacimiento),
        'BK': lambda e, ex: extraer_codigo_pais(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BL': lambda e, ex: extraer_codigo_ciudad(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BM': lambda e, ex: extraer_codigo_pais(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BN': lambda e, ex: extraer_codigo_ciudad(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BO': lambda e, ex: extraer_codigo_pais(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BP': lambda e, ex: extraer_codigo_ciudad(e.ciudad_residencia) if e.ciudad_residencia else '',
        'BQ': lambda e, ex: '0',
        'BR': lambda e, ex: '0',
        'BS': lambda e, ex: TIPO_COTIZANTE_NUM_MAP.get(e.tipo_cotizante, '0'),
        'BT': lambda e, ex: '0',
        'BU': lambda e, ex: es_extranjero_a_numero(e),          # mismo criterio que nacionalidad
        'BV': lambda e, ex: bool_a_01(e.reside_extranjero),
        'BW': lambda e, ex: '',
        'BX': lambda e, ex: '0',
        'BY': lambda e, ex: '',
        'BZ': lambda e, ex: str(e.horas_mes) if e.horas_mes is not None else '240',
        'CA': lambda e, ex: '1',                                # Número de contrato siempre 1
        'CB': lambda e, ex: '0',
        'CC': lambda e, ex: formato_fecha(e.f_retiro),
        'CD': lambda e, ex: '',
        'CE': lambda e, ex: '0',
        'CF': lambda e, ex: '0',
        'CG': lambda e, ex: '0',
        'CH': lambda e, ex: '0',
        'CI': lambda e, ex: '0',
        'CJ': lambda e, ex: '0',
        'CK': lambda e, ex: '0',
        'CL': lambda e, ex: '',
        'CM': lambda e, ex: '0',
        'CN': lambda e, ex: '0',
        'CO': lambda e, ex: '0',
        'CP': lambda e, ex: bool_a_01(e.deducible_medicina),
        'CQ': lambda e, ex: bool_a_01(e.deducible_vivienda),
        'CR': lambda e, ex: str(e.personas_acargo) if e.personas_acargo is not None else '',
        'CS': lambda e, ex: '0',
        'CT': lambda e, ex: '0',
        'CU': lambda e, ex: '0',
        'CV': lambda e, ex: '0',
        'CW': lambda e, ex: '0',
        'CX': lambda e, ex: '0',
        'CY': lambda e, ex: '0',
        'CZ': lambda e, ex: '0',
        'DA': lambda e, ex: '0',
        'DB': lambda e, ex: '0',
        'DC': lambda e, ex: '0',
        'DD': lambda e, ex: '0',
        'DE': lambda e, ex: '0',
        'DF': lambda e, ex: e.documento or '',
        'DG': lambda e, ex: '',
        'DH': lambda e, ex: '0',
        'DI': lambda e, ex: '0',
        'DJ': lambda e, ex: '0',
        'DK': lambda e, ex: '0',
        'DL': lambda e, ex: '0',
        'DM': lambda e, ex: '0',
        'DN': lambda e, ex: '0',
        'DO': lambda e, ex: '0',
        'DP': lambda e, ex: '0',
        'DQ': lambda e, ex: e.subcliente or '',
    }
    return MAPEO


def generar_excel_empleados(lista_empleados, ruta_plantilla, datos_extra_dict=None):
    if datos_extra_dict is None:
        datos_extra_dict = {}

    wb = load_workbook(ruta_plantilla)
    ws = wb['EMPLEADOS']

    FILA_EJEMPLO = 6
    max_col = ws.max_column

    # Colectar estilos de la fila ejemplo
    estilos_fila = []
    for col in range(1, max_col + 1):
        celda = ws.cell(row=FILA_EJEMPLO, column=col)
        estilos_fila.append({
            'font': copy(celda.font),
            'fill': copy(celda.fill),
            'border': copy(celda.border),
            'alignment': copy(celda.alignment),
            'number_format': celda.number_format,
        })

    # Limpiar contenido desde fila 6
    for row in ws.iter_rows(min_row=FILA_EJEMPLO, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.value = None

    if len(lista_empleados) > 1:
        ws.insert_rows(FILA_EJEMPLO + 1, amount=len(lista_empleados) - 1)

    mapeo = crear_mapeo_columnas(datos_extra_dict)

    for idx, empleado in enumerate(lista_empleados):
        fila_destino = FILA_EJEMPLO + idx
        for col_idx, estilo in enumerate(estilos_fila, start=1):
            celda_destino = ws.cell(row=fila_destino, column=col_idx)
            celda_destino.font = estilo['font']
            celda_destino.fill = estilo['fill']
            celda_destino.border = estilo['border']
            celda_destino.alignment = estilo['alignment']
            celda_destino.number_format = estilo['number_format']

        for col_letra, func in mapeo.items():
            col_idx = ws[f'{col_letra}1'].column
            valor = func(empleado, datos_extra_dict)
            ws.cell(row=fila_destino, column=col_idx, value=valor)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output